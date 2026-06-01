"""
EOL Agent — Excel Report Generator
Produces a professional 2-sheet workbook:
  Sheet 1 → At-Risk Devices  (EOL / EOS / expiring ≤24 months)
  Sheet 2 → Safe Devices      (3+ years remaining)
  Sheet 3 → Executive Summary (charts-ready stats)
"""

from datetime import date
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.styles.numbers import FORMAT_DATE_DDMMYY
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.image import Image
from core.models import AnalysisResult, EOLStatus


# ── Colour palette ─────────────────────────────────────────────── #
C_RED_FILL    = "FDECEA"
C_RED_FONT    = "9B1C1C"
C_AMBER_FILL  = "FFF8E1"
C_AMBER_FONT  = "7E5309"
C_YELLOW_FILL = "FFFDE7"
C_YELLOW_FONT = "5C4B00"
C_GREEN_FILL  = "E8F5E9"
C_GREEN_FONT  = "1B5E20"
C_BLUE_FILL   = "E3F2FD"
C_BLUE_FONT   = "0D47A1"
C_GREY_FILL   = "F5F5F5"
C_HEADER_BG   = "1E3A5F"
C_HEADER_FG   = "FFFFFF"
C_ALT_ROW     = "F9FAFB"
C_BORDER      = "D0D7DE"


STATUS_STYLE = {
    EOLStatus.EXPIRED:  (C_RED_FILL,    C_RED_FONT,    "EOL/EOS — Expired"),
    EOLStatus.CRITICAL: (C_RED_FILL,    C_RED_FONT,    "Critical — <6 months"),
    EOLStatus.WARNING:  (C_AMBER_FILL,  C_AMBER_FONT,  "Warning — 6–12 months"),
    EOLStatus.MONITOR:  (C_YELLOW_FILL, C_YELLOW_FONT, "Monitor — 12–24 months"),
    EOLStatus.SAFE:     (C_GREEN_FILL,  C_GREEN_FONT,  "Safe — 2+ years"),
    EOLStatus.UNKNOWN:  (C_GREY_FILL,   "555555",      "EOS/EOL Not Announced"),
}

MAX_ROWS_AT_RISK = 5000
MAX_ROWS_SAFE    = 3000


class ExcelReportGenerator:

    AT_RISK_COLUMNS = [
        ("Device Name",         20),
        ("Part Code",           22),
        ("Vendor",              12),
        ("Company",             20),
        ("Country",             14),
        ("Model",               28),
        ("Serial Number",       18),
        ("IP Address",          14),
        ("Location",            16),
        ("Assigned To",         16),
        ("Qty",                  5),
        ("End of Sale",         13),
        ("End of Support",      14),
        ("End of SW Maint.",    14),
        ("Days Remaining",      14),
        ("Status",              22),
        ("Replacement SKU",     22),
        ("Replacement Name",    28),
        ("Migration URL",       40),
        ("Source / Reference",  40),
        ("Notes",               50),
    ]

    SAFE_COLUMNS = [
        ("Device Name",         20),
        ("Part Code",           22),
        ("Vendor",              12),
        ("Company",             20),
        ("Country",             14),
        ("Model",               28),
        ("Serial Number",       18),
        ("IP Address",          14),
        ("Location",            16),
        ("Assigned To",         16),
        ("Qty",                  5),
        ("End of Sale",         13),
        ("End of Support",      14),
        ("Days Remaining",      14),
        ("Status",              22),
        ("Notes",               50),
    ]

    def __init__(self, output_path: str):
        self.output_path = output_path
        self.wb = Workbook()
        # Pre-create border once — reused for every cell (huge perf win)
        _side = Side(style="thin", color=C_BORDER)
        self._border = Border(left=_side, right=_side, top=_side, bottom=_side)

    # ─────────────────────────────────────────────────────────────── #

    def generate(
        self,
        at_risk: list[AnalysisResult],
        safe:    list[AnalysisResult],
        stats:   dict,
        generated_by: str = "EOL Agent",
    ) -> str:
        wb = self.wb
        wb.remove(wb.active)   # Remove default blank sheet

        ws_risk    = wb.create_sheet("⚠ At-Risk Devices")
        ws_safe    = wb.create_sheet("✅ Safe Devices")
        ws_summary = wb.create_sheet("📊 Executive Summary")

        self._build_sheet(ws_risk, at_risk, self.AT_RISK_COLUMNS, sheet_type="at_risk")
        self._build_sheet(ws_safe, safe,    self.SAFE_COLUMNS,    sheet_type="safe")
        self._build_summary(ws_summary, stats, generated_by)

        # Sheet tab colours
        ws_risk.sheet_properties.tabColor    = "C0392B"
        ws_safe.sheet_properties.tabColor    = "27AE60"
        ws_summary.sheet_properties.tabColor = "2980B9"

        wb.save(self.output_path)
        print(f"[Report] Saved -> {self.output_path}")
        return self.output_path

    # ─────────────────────────────────────────────────────────────── #
    #  Sheet builder                                                   #
    # ─────────────────────────────────────────────────────────────── #

    def _build_sheet(
        self,
        ws,
        results: list[AnalysisResult],
        columns: list[tuple],
        sheet_type: str,
    ):
        max_rows = MAX_ROWS_AT_RISK if sheet_type == "at_risk" else MAX_ROWS_SAFE
        truncated = len(results) > max_rows
        results = results[:max_rows]

        # ── Title row ── #
        title = "At-Risk Network Devices — EOL / EOS Report" if sheet_type == "at_risk" \
                else "Safe Network Devices — Lifecycle Status"
        last_col = get_column_letter(len(columns))
        ws.merge_cells(f"A1:{last_col}1")
        title_cell = ws["A1"]
        title_cell.value     = title
        title_cell.font      = Font(name="Arial", bold=True, size=14, color=C_HEADER_FG)
        title_cell.fill      = PatternFill("solid", fgColor=C_HEADER_BG)
        title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[1].height = 32

        # ── Sub-header (metadata) ── #
        ws.merge_cells(f"A2:{last_col}2")
        meta = ws["A2"]
        trunc_note = f"  (showing top {max_rows} of {len(results)} total)" if truncated else ""
        meta.value     = f"Generated: {date.today().strftime('%d %B %Y')}   |   Records: {len(results)}{trunc_note}"
        meta.font      = Font(name="Arial", size=10, color="888888")
        meta.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[2].height = 18

        # ── Column headers ── #
        hdr_fill  = PatternFill("solid", fgColor="2C3E50")
        hdr_font  = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border    = self._border
        for col_idx, (col_name, col_width) in enumerate(columns, 1):
            cell = ws.cell(row=3, column=col_idx, value=col_name)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = hdr_align
            cell.border    = border
            ws.column_dimensions[get_column_letter(col_idx)].width = col_width
        ws.row_dimensions[3].height = 22

        # ── Pre-create all reusable style objects (major perf win) ── #
        font_base    = Font(name="Arial", size=9)
        fill_white   = PatternFill("solid", fgColor="FFFFFF")
        fill_alt     = PatternFill("solid", fgColor=C_ALT_ROW)
        align_base   = Alignment(vertical="center", wrap_text=False)
        align_center = Alignment(horizontal="center", vertical="center")

        # Status styles per EOL category
        _st_fills  = {k: PatternFill("solid", fgColor=v[0]) for k, v in STATUS_STYLE.items()}
        _st_fonts  = {k: Font(name="Arial", size=9, bold=True, color=v[1]) for k, v in STATUS_STYLE.items()}
        _fallback_fill = PatternFill("solid", fgColor=C_GREY_FILL)
        _fallback_font = Font(name="Arial", size=9, bold=True, color="555555")

        # Days-remaining fonts
        font_red   = Font(name="Arial", size=9, bold=True, color=C_RED_FONT)
        font_amber = Font(name="Arial", size=9, bold=True, color=C_AMBER_FONT)
        font_yel   = Font(name="Arial", size=9, color=C_YELLOW_FONT)

        status_col = 16 if sheet_type == "at_risk" else 15
        days_col   = 15  # only relevant for at_risk sheet

        def fmt_date(dt) -> str:
            return dt.strftime("%d %b %Y") if dt else "-"

        # ── Data rows ── #
        for row_idx, result in enumerate(results, 4):
            d = result.device
            e = result.eol
            is_alt = (row_idx % 2 == 0)
            base_fill = fill_alt if is_alt else fill_white

            sfill = _st_fills.get(result.status, _fallback_fill)
            sfont = _st_fonts.get(result.status, _fallback_font)
            sl, _, status_label = STATUS_STYLE.get(result.status, (C_GREY_FILL, "555555", "EOS/EOL Not Announced"))

            _loc_parts = (d.location or "").split(" | ")
            _company = _loc_parts[0].strip() if _loc_parts else "-"
            _country = _loc_parts[-1].strip() if len(_loc_parts) > 1 else "-"

            if sheet_type == "at_risk":
                row_data = [
                    d.name, d.part_code, d.vendor, _company, _country, d.model,
                    d.serial or "-", d.ip_address or "-", d.location or "-", d.assigned_to or "-",
                    d.quantity,
                    fmt_date(e.end_of_sale)     if e else "-",
                    fmt_date(e.end_of_support)  if e else "-",
                    fmt_date(e.end_of_sw_maint) if e else "-",
                    result.days_left if result.days_left is not None else "-",
                    status_label,
                    (e.replacement_sku  or "-") if e else "-",
                    (e.replacement_name or "-") if e else "-",
                    (e.migration_url    or "-") if e else "-",
                    (e.source_url       or "-") if e else "-",
                    result.notes or "-",
                ]
            else:
                row_data = [
                    d.name, d.part_code, d.vendor, _company, _country, d.model,
                    d.serial or "-", d.ip_address or "-", d.location or "-", d.assigned_to or "-",
                    d.quantity,
                    fmt_date(e.end_of_sale)    if e else "-",
                    fmt_date(e.end_of_support) if e else "-",
                    result.days_left if result.days_left is not None else "-",
                    status_label,
                    result.notes or "-",
                ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                if col_idx == status_col:
                    cell.fill      = sfill
                    cell.font      = sfont
                    cell.alignment = align_center
                elif col_idx == days_col and sheet_type == "at_risk" and isinstance(value, int):
                    cell.fill      = base_fill
                    cell.alignment = align_center
                    if value < 180:
                        cell.font = font_red
                    elif value < 365:
                        cell.font = font_amber
                    else:
                        cell.font = font_yel
                else:
                    cell.fill      = base_fill
                    cell.font      = font_base
                    cell.alignment = align_base

        # ── Freeze panes & auto-filter ── #
        ws.freeze_panes = "A4"
        ws.auto_filter.ref = f"A3:{last_col}{3 + len(results)}"

    # ─────────────────────────────────────────────────────────────── #
    #  Executive Summary sheet                                         #
    # ─────────────────────────────────────────────────────────────── #

    def _build_summary(self, ws, stats: dict, generated_by: str):
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 18

        # Title
        ws.merge_cells("A1:E1")
        c = ws["A1"]
        c.value     = "📊  Network Device EOL — Executive Summary"
        c.font      = Font(name="Arial", bold=True, size=14, color=C_HEADER_FG)
        c.fill      = PatternFill("solid", fgColor=C_HEADER_BG)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[1].height = 32

        ws.merge_cells("A2:E2")
        ws["A2"].value = f"Report generated: {date.today().strftime('%d %B %Y')}   |   By: {generated_by}"
        ws["A2"].font  = Font(name="Arial", size=10, color="888888")
        ws.row_dimensions[2].height = 18

        # ── KPI table ── #
        kpi_rows = [
            ("Metric",                    "Count",  "% of Total", "Status"),
            ("Total Devices Analysed",    stats["total"],    "=B5/B5",         "—"),
            ("EOL/EOS — Expired",         stats["expired"],  f'=B6/$B$5',      "🔴 Immediate Action"),
            ("Critical — <6 months",      stats["critical"], f'=B7/$B$5',      "🔴 Start Procurement"),
            ("Warning — 6–12 months",     stats["warning"],  f'=B8/$B$5',      "🟡 Plan Replacement"),
            ("Monitor — 12–24 months",    stats["monitor"],  f'=B9/$B$5',      "🟠 Next Refresh Cycle"),
            ("Safe — 2+ years",           stats["safe"],     f'=B10/$B$5',     "🟢 No Action Required"),
            ("EOS/EOL Not Announced",     stats["unknown"],  f'=B11/$B$5',     "⚪ Verify Manually"),
        ]

        row_colours = [
            (C_HEADER_BG, C_HEADER_FG),
            ("ECEFF1",    "263238"),
            (C_RED_FILL,  C_RED_FONT),
            (C_RED_FILL,  C_RED_FONT),
            (C_AMBER_FILL, C_AMBER_FONT),
            (C_YELLOW_FILL, C_YELLOW_FONT),
            (C_GREEN_FILL,  C_GREEN_FONT),
            (C_GREY_FILL,   "555555"),
        ]

        for i, (row_data, (bg, fg)) in enumerate(zip(kpi_rows, row_colours), 4):
            for j, val in enumerate(row_data, 1):
                cell = ws.cell(row=i, column=j, value=val)
                cell.fill      = PatternFill("solid", fgColor=bg)
                cell.font      = Font(name="Arial", bold=(i == 4), size=10, color=fg)
                cell.alignment = Alignment(
                    horizontal="center" if j > 1 else "left",
                    vertical="center",
                    indent=1 if j == 1 else 0,
                )
                cell.border = self._thin_border()
                if j == 3 and i > 4:
                    cell.number_format = "0.0%"
            ws.row_dimensions[i].height = 20

        # ── Risk breakdown header ── #
        ws.cell(row=13, column=1, value="Risk Distribution Data (for charting)").font = \
            Font(name="Arial", bold=True, size=10)
        ws.row_dimensions[13].height = 18

        chart_data = [
            ("Category",          "Devices"),
            ("Expired",           stats["expired"]),
            ("Critical <6mo",     stats["critical"]),
            ("Warning 6–12mo",    stats["warning"]),
            ("Monitor 12–24mo",   stats["monitor"]),
            ("Safe 2+ yrs",       stats["safe"]),
            ("Not Announced",     stats["unknown"]),
        ]
        for i, (cat, val) in enumerate(chart_data, 14):
            ws.cell(row=i, column=1, value=cat).font = Font(name="Arial", size=9)
            ws.cell(row=i, column=2, value=val).font  = Font(name="Arial", size=9)

        # ── Bar chart ── #
        chart = BarChart()
        chart.type    = "col"
        chart.title   = "Device EOL Risk Distribution"
        chart.y_axis.title = "Number of Devices"
        chart.x_axis.title = "Risk Category"
        chart.style  = 10
        chart.width  = 20
        chart.height = 12

        data = Reference(ws, min_col=2, min_row=14, max_row=20)
        cats = Reference(ws, min_col=1, min_row=15, max_row=20)
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)

        # Colour each bar per risk level
        bar_colours = ["C0392B", "E74C3C", "F39C12", "F1C40F", "27AE60", "95A5A6"]
        for idx, colour in enumerate(bar_colours):
            pt = DataPoint(idx=idx)
            pt.spPr.solidFill = colour
            chart.series[0].dPt.append(pt)

        ws.add_chart(chart, "D4")

        # ── Footer ── #
        ws.cell(row=22, column=1, value="Notes:").font = Font(name="Arial", bold=True, size=9)
        notes = [
            "• End-of-Sale (EOS): The product is no longer available for purchase. Existing units continue to receive support.",
            "• End-of-Support (EOL): No further software updates, bug fixes, or TAC support. Security vulnerabilities will not be patched.",
            "• Replacement SKUs are vendor-recommended migration paths. Verify with your account team before ordering.",
            "• Devices marked 'EOS/EOL Not Announced' could not be matched to a known EOL record — verify manually with the vendor.",
        ]
        for i, note in enumerate(notes, 23):
            cell = ws.cell(row=i, column=1, value=note)
            cell.font      = Font(name="Arial", size=9, color="555555")
            ws.merge_cells(f"A{i}:E{i}")

    # ─────────────────────────────────────────────────────────────── #
    #  Helpers                                                         #
    # ─────────────────────────────────────────────────────────────── #

    def _thin_border(self) -> Border:
        return self._border
